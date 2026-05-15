from __future__ import annotations

import html
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Callable

import requests

from tool_protocol import ToolCall, first_line_and_body, parse_key_value_block, parse_replace_payload


ApprovalCallback = Callable[[str, str, bool, dict], tuple[bool, str]]


@dataclass(frozen=True)
class GatewayResult:
    ok: bool
    text: str


class ToolGateway:
    def __init__(self, config: dict, project_dir: str | Path, approval_callback: ApprovalCallback | None = None):
        self.config = config
        self.project_dir = Path(project_dir)
        self.data_dir = self.project_dir / "data"
        self.memory_path = self.data_dir / "memory" / "memory.json"
        self.memory_path.parent.mkdir(parents=True, exist_ok=True)
        self.approval_callback = approval_callback
        self.last_browser_url = ""
        self._memory_cache: dict | None = None
        self._memory_cache_mtime: float = 0.0

    def execute(self, call: ToolCall) -> GatewayResult:
        enabled = set(self.config.get("tools", {}).get("enabled", []))
        if call.name not in enabled:
            return GatewayResult(False, f"[ERROR] Tool '{call.name}' is disabled.")
        try:
            method = getattr(self, f"_tool_{call.name}", None)
            if not method:
                return GatewayResult(False, f"[ERROR] Tool '{call.name}' is not implemented.")
            return method(call.raw)
        except Exception as exc:
            return GatewayResult(False, f"[ERROR] Tool '{call.name}' failed: {exc}")

    def _approval_required(self, tool_name: str) -> bool:
        policy = str(self.config.get("app", {}).get("approval_policy", "guarded"))
        if policy == "full_auto":
            return False
        required = set(self.config.get("tools", {}).get("approval_required", []))
        if policy == "mostly_auto":
            return tool_name in {"write", "replace", "git_commit"}
        return tool_name in required

    def _ask_approval(self, title: str, content: str, editable: bool = True, metadata: dict | None = None) -> tuple[bool, str]:
        if not self.approval_callback:
            return False, content
        return self.approval_callback(title, content, editable, metadata or {})

    def _maybe_approve(self, tool_name: str, title: str, content: str, editable: bool = True, metadata: dict | None = None) -> tuple[bool, str]:
        if not self._approval_required(tool_name):
            return True, content
        return self._ask_approval(title, content, editable, metadata)

    def _allowed_roots(self) -> list[Path]:
        roots = []
        for item in self.config.get("security", {}).get("filesystem_roots", []):
            value = str(item).replace("$PROJECT_DIR", str(self.project_dir))
            roots.append(Path(value).expanduser().resolve())
        roots.append(self.project_dir.resolve())
        return roots

    def _resolve_allowed_path(self, raw_path: str, must_exist: bool = False, access_reason: str = "path access") -> Path:
        value = str(raw_path or "").strip().strip('"').strip("'")
        if not value:
            raise ValueError("path is empty")
        path = Path(value)
        if not path.is_absolute():
            path = self.project_dir / path
        resolved = path.resolve()
        if must_exist and not resolved.exists():
            raise FileNotFoundError(str(resolved))
        access_mode = str(self.config.get("security", {}).get("access_mode", "restricted")).lower()
        if access_mode == "full":
            return resolved
        allowed = False
        for root in self._allowed_roots():
            try:
                resolved.relative_to(root)
                allowed = True
                break
            except ValueError:
                continue
        if not allowed:
            if access_mode == "ask":
                approved, _ = self._ask_approval(
                    "Gemini path access request",
                    str(resolved),
                    False,
                    {"reason": access_reason, "access_mode": access_mode},
                )
                if approved:
                    return resolved
            raise PermissionError(f"path is outside allowed roots: {resolved}")
        return resolved

    def _read_text_limited(self, path: Path) -> str:
        max_bytes = int(self.config.get("security", {}).get("max_read_bytes", 200000))
        data = path.read_bytes()[:max_bytes]
        suffix = ""
        if path.stat().st_size > max_bytes:
            suffix = f"\n\n[TRUNCATED at {max_bytes} bytes]"
        return data.decode("utf-8", errors="replace") + suffix

    def _tool_read(self, raw: str) -> GatewayResult:
        path = self._resolve_allowed_path(raw, must_exist=True)
        if not path.is_file():
            return GatewayResult(False, f"[ERROR] Not a file: {path}")
        content = self._read_text_limited(path)
        return GatewayResult(True, f"[FILE {path}]\n{content}")

    def _tool_explorer(self, raw: str) -> GatewayResult:
        path = self._resolve_allowed_path(raw or ".", must_exist=True)
        if not path.is_dir():
            return GatewayResult(False, f"[ERROR] Not a directory: {path}")
        lines = [f"[EXPLORER {path}]"]
        for item in sorted(path.iterdir(), key=lambda p: (not p.is_dir(), p.name.lower()))[:300]:
            kind = "DIR" if item.is_dir() else "FILE"
            size = "" if item.is_dir() else f" {item.stat().st_size} bytes"
            lines.append(f"- [{kind}] {item.name}{size}")
        return GatewayResult(True, "\n".join(lines))

    def _tool_list_dir(self, raw: str) -> GatewayResult:
        return self._tool_explorer(raw)

    def _tool_file_info(self, raw: str) -> GatewayResult:
        path = self._resolve_allowed_path(raw, must_exist=True)
        stat = path.stat()
        payload = {
            "path": str(path),
            "exists": path.exists(),
            "is_file": path.is_file(),
            "is_dir": path.is_dir(),
            "size": stat.st_size,
            "modified": stat.st_mtime,
        }
        return GatewayResult(True, json.dumps(payload, indent=2))

    def _check_placeholders(self, text: str) -> str:
        warning = ""
        placeholders = ["[Insert", "TODO:", "...", "// implementation", "[add ", "[your ", "[...]"]
        if any(p in text for p in placeholders):
            warning = "\n[WARNING] File written, but appears to contain placeholder text. Please emit full actual contents."
        return warning

    def _tool_write(self, raw: str) -> GatewayResult:
        path_text, body = first_line_and_body(raw)
        path = self._resolve_allowed_path(path_text, must_exist=False, access_reason="write file")

        warning = self._check_placeholders(body)
        meta = {"path": str(path)}
        if warning:
            meta["warning"] = warning.strip()

        approved, edited = self._maybe_approve("write", "Gemini write request", body, True, meta)
        if not approved:
            return GatewayResult(False, "[BLOCKED] User denied file write.")
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(edited, encoding="utf-8")

        if not path.exists():
            return GatewayResult(False, f"[ERROR] Failed to verify file existence after write: {path}")

        final_warning = self._check_placeholders(edited)
        return GatewayResult(True, f"[OK] Wrote {edited.count(chr(10)) + 1} lines to {path}{final_warning}")

    def _tool_replace(self, raw: str) -> GatewayResult:
        path_text, old, new = parse_replace_payload(raw)
        path = self._resolve_allowed_path(path_text, must_exist=True, access_reason="replace file content")
        current = path.read_text(encoding="utf-8", errors="replace")
        count = current.count(old)
        if count == 0:
            return GatewayResult(False, "[ERROR] Old snippet was not found.")
        if count > 1:
            return GatewayResult(False, f"[ERROR] Old snippet matched {count} times; use a larger exact snippet.")
        updated = current.replace(old, new, 1)

        warning = self._check_placeholders(updated)
        meta = {"path": str(path)}
        if warning:
            meta["warning"] = warning.strip()

        approved, edited = self._maybe_approve("replace", "Gemini replace request", updated, True, meta)
        if not approved:
            return GatewayResult(False, "[BLOCKED] User denied file replacement.")
        path.write_text(edited, encoding="utf-8")

        if not path.exists():
            return GatewayResult(False, f"[ERROR] Failed to verify file existence after replace: {path}")

        final_warning = self._check_placeholders(edited)
        return GatewayResult(True, f"[OK] Replaced snippet in {path}{final_warning}")

    def _tool_append(self, raw: str) -> GatewayResult:
        path_text, body = first_line_and_body(raw)
        path = self._resolve_allowed_path(path_text, must_exist=False, access_reason="append to file")

        warning = self._check_placeholders(body)
        meta = {"path": str(path)}
        if warning:
            meta["warning"] = warning.strip()

        approved, edited = self._maybe_approve("append", "Gemini append request", body, True, meta)
        if not approved:
            return GatewayResult(False, "[BLOCKED] User denied file append.")
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("a", encoding="utf-8") as handle:
            handle.write(edited)

        if not path.exists():
            return GatewayResult(False, f"[ERROR] Failed to verify file existence after append: {path}")

        final_warning = self._check_placeholders(edited)
        return GatewayResult(True, f"[OK] Appended {len(edited)} characters to {path}{final_warning}")

    def _tool_mkdir(self, raw: str) -> GatewayResult:
        path = self._resolve_allowed_path(raw, must_exist=False, access_reason="create directory")
        approved, _ = self._maybe_approve("mkdir", "Gemini create folder request", str(path), False, {"path": str(path)})
        if not approved:
            return GatewayResult(False, "[BLOCKED] User denied folder creation.")
        path.mkdir(parents=True, exist_ok=True)
        return GatewayResult(True, f"[OK] Created directory: {path}")

    def _tool_copy(self, raw: str) -> GatewayResult:
        source_text, dest_text = first_line_and_body(raw)
        source = self._resolve_allowed_path(source_text, must_exist=True, access_reason="copy source")
        dest = self._resolve_allowed_path(dest_text.strip(), must_exist=False, access_reason="copy destination")
        approved, _ = self._maybe_approve("copy", "Gemini copy request", f"{source}\n-> {dest}", False, {"source": str(source), "destination": str(dest)})
        if not approved:
            return GatewayResult(False, "[BLOCKED] User denied copy.")
        if source.is_dir():
            shutil.copytree(source, dest, dirs_exist_ok=True)
        else:
            dest.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source, dest)
        return GatewayResult(True, f"[OK] Copied {source} to {dest}")

    def _tool_move(self, raw: str) -> GatewayResult:
        source_text, dest_text = first_line_and_body(raw)
        source = self._resolve_allowed_path(source_text, must_exist=True, access_reason="move source")
        dest = self._resolve_allowed_path(dest_text.strip(), must_exist=False, access_reason="move destination")
        approved, _ = self._maybe_approve("move", "Gemini move request", f"{source}\n-> {dest}", False, {"source": str(source), "destination": str(dest)})
        if not approved:
            return GatewayResult(False, "[BLOCKED] User denied move.")
        dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.move(str(source), str(dest))
        return GatewayResult(True, f"[OK] Moved {source} to {dest}")

    def _tool_delete(self, raw: str) -> GatewayResult:
        data = parse_key_value_block(raw)
        path_text = data.get("path") or str(raw or "").splitlines()[0].strip()
        recursive = str(data.get("recursive", "false")).lower() in {"1", "true", "yes", "y"}
        path = self._resolve_allowed_path(path_text, must_exist=True, access_reason="delete path")
        approved, _ = self._maybe_approve("delete", "Gemini delete request", str(path), False, {"path": str(path), "recursive": recursive})
        if not approved:
            return GatewayResult(False, "[BLOCKED] User denied delete.")
        if path.is_dir():
            if recursive:
                shutil.rmtree(path)
            else:
                path.rmdir()
        else:
            path.unlink()
        return GatewayResult(True, f"[OK] Deleted {path}")

    def _tool_terminal(self, raw: str) -> GatewayResult:
        command = str(raw or "").strip()
        if not command:
            return GatewayResult(False, "[ERROR] terminal command is empty.")
        approved, edited = self._maybe_approve("terminal", "Gemini terminal command", command, True, {"cwd": self._terminal_cwd()})
        if not approved:
            return GatewayResult(False, "[BLOCKED] User denied terminal command.")
        timeout = int(self.config.get("security", {}).get("terminal_timeout_seconds", 30))
        completed = subprocess.run(
            edited,
            shell=True,
            cwd=self._terminal_cwd(),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=timeout,
        )
        output = (completed.stdout or "").strip()
        if completed.stderr:
            output += f"\n[STDERR]\n{completed.stderr.strip()}"
        if completed.returncode != 0:
            return GatewayResult(False, f"[COMMAND EXIT {completed.returncode}]\n{output}")
        return GatewayResult(True, output or "[OK] Command completed with no output.")

    def _tool_powershell(self, raw: str) -> GatewayResult:
        command = str(raw or "").strip()
        if not command:
            return GatewayResult(False, "[ERROR] PowerShell command is empty.")
        approved, edited = self._maybe_approve("powershell", "Gemini PowerShell command", command, True, {"cwd": self._terminal_cwd()})
        if not approved:
            return GatewayResult(False, "[BLOCKED] User denied PowerShell command.")
        return self._run_process(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", edited],
            cwd=self._terminal_cwd(),
            timeout=int(self.config.get("security", {}).get("terminal_timeout_seconds", 30)),
            label="POWERSHELL",
        )

    def _tool_wsl(self, raw: str) -> GatewayResult:
        command = str(raw or "").strip()
        if not command:
            return GatewayResult(False, "[ERROR] WSL command is empty.")
        approved, edited = self._maybe_approve("wsl", "Gemini WSL/Linux command", command, True, {"cwd": self._terminal_cwd()})
        if not approved:
            return GatewayResult(False, "[BLOCKED] User denied WSL command.")
        return self._run_process(
            ["wsl", "bash", "-lc", edited],
            cwd=self._terminal_cwd(),
            timeout=int(self.config.get("security", {}).get("terminal_timeout_seconds", 30)),
            label="WSL",
        )

    def _run_process(self, command: list[str], cwd: str | None = None, timeout: int = 30, label: str = "PROCESS") -> GatewayResult:
        completed = subprocess.run(
            command,
            cwd=cwd,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=timeout,
        )
        output = (completed.stdout or "").strip()
        if completed.stderr:
            output += f"\n[STDERR]\n{completed.stderr.strip()}"
        if completed.returncode != 0:
            return GatewayResult(False, f"[{label} EXIT {completed.returncode}]\n{output}")
        return GatewayResult(True, output or f"[OK] {label} command completed with no output.")

    def _terminal_cwd(self) -> str:
        return str(Path(self.config.get("security", {}).get("terminal_cwd") or self.project_dir).resolve())

    def _tool_fetch(self, raw: str) -> GatewayResult:
        url = str(raw or "").strip()
        if not re.match(r"(?i)^https?://", url):
            return GatewayResult(False, "[ERROR] fetch requires an absolute http:// or https:// URL.")
        session = requests.Session()
        session.trust_env = False
        response = session.get(url, timeout=25, headers={"User-Agent": "GeminiLocalAgent/1.0"})
        response.raise_for_status()
        text = response.text
        if "text/html" in response.headers.get("content-type", ""):
            text = self._strip_html(text)
        self.last_browser_url = url
        return GatewayResult(True, f"[FETCH {url}]\n{text[:50000]}")

    def _tool_browser_open(self, raw: str) -> GatewayResult:
        result = self._tool_fetch(raw)
        if result.ok:
            self.last_browser_url = str(raw or "").strip()
        return result

    def _tool_browser_inspect(self, raw: str) -> GatewayResult:
        target = str(raw or "").strip()
        if not target or target.lower() == "current":
            target = self.last_browser_url
        if not target:
            return GatewayResult(False, "[ERROR] No browser URL has been opened yet.")
        return self._tool_fetch(target)

    def _tool_browser_screenshot(self, raw: str) -> GatewayResult:
        target = str(raw or self.last_browser_url or "").strip()
        approved, _ = self._maybe_approve("browser_screenshot", "Gemini browser screenshot", target, False, {"url": target})
        if not approved:
            return GatewayResult(False, "[BLOCKED] User denied browser screenshot.")
        if not re.match(r"(?i)^https?://", target):
            return GatewayResult(False, "[ERROR] browser_screenshot requires an absolute http:// or https:// URL.")
        out_dir = self.data_dir / "logs" / "screenshots"
        out_dir.mkdir(parents=True, exist_ok=True)
        safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", target)[:80].strip("_") or "page"
        out_path = out_dir / f"{safe_name}.png"
        command = [
            "npx",
            "-y",
            "playwright",
            "screenshot",
            "--wait-for-timeout=1000",
            target,
            str(out_path),
        ]
        completed = subprocess.run(
            command,
            cwd=str(self.project_dir),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=90,
        )
        output = (completed.stdout or "").strip()
        if completed.stderr:
            output += f"\n[STDERR]\n{completed.stderr.strip()}"
        if completed.returncode != 0:
            return GatewayResult(False, f"[PLAYWRIGHT EXIT {completed.returncode}]\n{output}")
        return GatewayResult(True, f"[OK] Browser screenshot saved: {out_path}")

    def _tool_browser_click(self, raw: str) -> GatewayResult:
        return self._run_playwright_action("click", raw)

    def _tool_browser_type(self, raw: str) -> GatewayResult:
        return self._run_playwright_action("type", raw)

    def _run_playwright_action(self, action: str, raw: str) -> GatewayResult:
        data = parse_key_value_block(raw)
        url = data.get("url", "").strip() or self.last_browser_url
        selector = data.get("selector", "").strip()
        text = data.get("text", "")
        if not re.match(r"(?i)^https?://", url or ""):
            return GatewayResult(False, "[ERROR] Browser action requires a URL.")
        if not selector:
            return GatewayResult(False, "[ERROR] Browser action requires a selector.")
        if action == "type" and not text:
            return GatewayResult(False, "[ERROR] browser_type requires text.")
        approved, _ = self._maybe_approve(
            f"browser_{action}",
            f"Gemini browser {action}",
            raw,
            True,
            {"url": url, "selector": selector},
        )
        if not approved:
            return GatewayResult(False, f"[BLOCKED] User denied browser {action}.")
        script = self._playwright_script(action, url, selector, text)
        with tempfile.NamedTemporaryFile("w", suffix=".py", delete=False, encoding="utf-8") as handle:
            handle.write(script)
            script_path = handle.name
        try:
            result = self._run_process([sys.executable, script_path], cwd=str(self.project_dir), timeout=90, label="PLAYWRIGHT")
            if result.ok:
                self.last_browser_url = url
            return result
        finally:
            try:
                os.unlink(script_path)
            except OSError:
                pass

    @staticmethod
    def _playwright_script(action: str, url: str, selector: str, text: str) -> str:
        return f"""
import json
from playwright.sync_api import sync_playwright

url = {json.dumps(url)}
selector = {json.dumps(selector)}
text = {json.dumps(text)}
action = {json.dumps(action)}

with sync_playwright() as p:
    browser = p.chromium.launch(headless=True)
    page = browser.new_page()
    page.goto(url, wait_until="domcontentloaded", timeout=45000)
    page.wait_for_selector(selector, timeout=15000)
    if action == "click":
        page.click(selector)
    else:
        page.fill(selector, text)
    print(json.dumps({{"ok": True, "url": page.url, "title": page.title()}}))
    browser.close()
"""

    def _tool_git_status(self, raw: str) -> GatewayResult:
        return self._run_git(raw, ["status", "--short"], approval=False)

    def _tool_git_diff(self, raw: str) -> GatewayResult:
        return self._run_git(raw, ["diff", "--"], approval=False)

    def _tool_git_add(self, raw: str) -> GatewayResult:
        repo, body = first_line_and_body(raw)
        pathspecs = [line.strip() for line in body.splitlines() if line.strip()] or ["."]
        approved, edited = self._maybe_approve("git_add", "Gemini git add", "\n".join(pathspecs), True, {"repo": repo})
        if not approved:
            return GatewayResult(False, "[BLOCKED] User denied git add.")
        return self._run_git(repo, ["add", "--", *[line.strip() for line in edited.splitlines() if line.strip()]], approval=False)

    def _tool_git_commit(self, raw: str) -> GatewayResult:
        repo, message = first_line_and_body(raw)
        message = message.strip()
        if not message:
            return GatewayResult(False, "[ERROR] git_commit requires a commit message.")
        approved, edited = self._maybe_approve("git_commit", "Gemini git commit", message, True, {"repo": repo})
        if not approved:
            return GatewayResult(False, "[BLOCKED] User denied git commit.")
        return self._run_git(repo, ["commit", "-m", edited.strip()], approval=False)

    def _run_git(self, repo_raw: str, args: list[str], approval: bool = False) -> GatewayResult:
        repo = self._resolve_allowed_path(repo_raw or self.project_dir, must_exist=True)
        if not repo.is_dir():
            repo = repo.parent
        completed = subprocess.run(
            ["git", *args],
            cwd=str(repo),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=30,
        )
        output = (completed.stdout or "").strip()
        if completed.stderr:
            output += f"\n[STDERR]\n{completed.stderr.strip()}"
        if completed.returncode != 0:
            return GatewayResult(False, f"[GIT EXIT {completed.returncode}]\n{output}")
        return GatewayResult(True, output or "[OK] Git command completed with no output.")

    def _tool_memory_get(self, raw: str) -> GatewayResult:
        key = str(raw or "").strip()
        memory = self._load_memory()
        if key:
            return GatewayResult(True, json.dumps({key: memory.get(key, "")}, indent=2, ensure_ascii=False))
        return GatewayResult(True, json.dumps(memory, indent=2, ensure_ascii=False))

    def _tool_memory_put(self, raw: str) -> GatewayResult:
        key, value = first_line_and_body(raw)
        key = key.strip()
        if not key:
            data = parse_key_value_block(raw)
            key = data.get("key", "").strip()
            value = data.get("value", "")
        if not key:
            return GatewayResult(False, "[ERROR] memory_put requires a key on the first line.")
        memory = self._load_memory()
        memory[key] = value.strip()
        self.memory_path.write_text(json.dumps(memory, indent=2, ensure_ascii=False), encoding="utf-8")
        self._memory_cache = memory
        self._memory_cache_mtime = self.memory_path.stat().st_mtime
        return GatewayResult(True, f"[OK] Stored memory key: {key}")

    def _tool_mcp_status(self, raw: str) -> GatewayResult:
        lines = ["[MCP STATUS] Normalized Gemini tools are active."]
        for name, spec in sorted((self.config.get("mcp_servers") or {}).items()):
            enabled = bool(spec.get("enabled"))
            command = spec.get("command") or spec.get("url") or "[configured]"
            lines.append(f"- {name}: {'enabled' if enabled else 'disabled'} -> {command}")
        if not self.config.get("mcp_servers", {}).get("github", {}).get("enabled"):
            lines.append("- github: disabled by default")
        return GatewayResult(True, "\n".join(lines))

    def _tool_extract_text(self, raw: str) -> GatewayResult:
        path = self._resolve_allowed_path(raw, must_exist=True)
        markitdown = (self.config.get("mcp_servers", {}).get("markitdown") or {}).get("command") or ""
        if markitdown:
            try:
                completed = subprocess.run(
                    [markitdown, str(path)],
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    timeout=60,
                )
                if completed.returncode == 0 and completed.stdout.strip():
                    return GatewayResult(True, completed.stdout[:100000])
            except Exception:
                pass
        if path.suffix.lower() in {".txt", ".md", ".csv", ".json", ".py", ".js", ".html", ".css"}:
            return self._tool_read(str(path))
        return GatewayResult(False, "[ERROR] MarkItDown is unavailable and this file type has no fallback extractor.")

    def _tool_window_list(self, raw: str) -> GatewayResult:
        script = (
            "Get-Process | Where-Object {$_.MainWindowTitle} | "
            "Select-Object Id,ProcessName,MainWindowTitle | ConvertTo-Json -Depth 2"
        )
        return self._run_process(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
            cwd=self._terminal_cwd(),
            timeout=15,
            label="WINDOW_LIST",
        )

    def _tool_window_focus(self, raw: str) -> GatewayResult:
        target = str(raw or "").strip()
        if not target:
            return GatewayResult(False, "[ERROR] window_focus requires a process id or title fragment.")
        approved, _ = self._maybe_approve("window_focus", "Gemini window focus request", target, False, {"target": target})
        if not approved:
            return GatewayResult(False, "[BLOCKED] User denied window focus.")
        target_json = json.dumps(target)
        script = f"""
$target = {target_json}
$sig = '[DllImport("user32.dll")] public static extern bool SetForegroundWindow(IntPtr hWnd);'
Add-Type -MemberDefinition $sig -Name NativeMethods -Namespace Win32
$proc = $null
if ($target -match '^\\d+$') {{
  $proc = Get-Process -Id ([int]$target) -ErrorAction SilentlyContinue
}} else {{
  $proc = Get-Process | Where-Object {{ $_.MainWindowTitle -like "*$target*" }} | Select-Object -First 1
}}
if (-not $proc) {{ throw "Window not found: $target" }}
[Win32.NativeMethods]::SetForegroundWindow($proc.MainWindowHandle) | Out-Null
"Focused: $($proc.ProcessName) [$($proc.Id)] $($proc.MainWindowTitle)"
"""
        return self._run_process(["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script], cwd=self._terminal_cwd(), timeout=20, label="WINDOW_FOCUS")

    def _tool_window_close(self, raw: str) -> GatewayResult:
        target = str(raw or "").strip()
        if not target:
            return GatewayResult(False, "[ERROR] window_close requires a process id or title fragment.")
        approved, _ = self._maybe_approve("window_close", "Gemini window close request", target, False, {"target": target})
        if not approved:
            return GatewayResult(False, "[BLOCKED] User denied window close.")
        target_json = json.dumps(target)
        script = f"""
$target = {target_json}
$proc = $null
if ($target -match '^\\d+$') {{
  $proc = Get-Process -Id ([int]$target) -ErrorAction SilentlyContinue
}} else {{
  $proc = Get-Process | Where-Object {{ $_.MainWindowTitle -like "*$target*" }} | Select-Object -First 1
}}
if (-not $proc) {{ throw "Window not found: $target" }}
$ok = $proc.CloseMainWindow()
"Close requested: $ok for $($proc.ProcessName) [$($proc.Id)] $($proc.MainWindowTitle)"
"""
        return self._run_process(["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script], cwd=self._terminal_cwd(), timeout=20, label="WINDOW_CLOSE")

    def _tool_gui_click(self, raw: str) -> GatewayResult:
        data = parse_key_value_block(raw)
        coords = data.get("point") or data.get("xy") or str(raw or "")
        match = re.search(r"(-?\d+)\s*,\s*(-?\d+)", coords)
        if not match:
            return GatewayResult(False, "[ERROR] gui_click requires x,y coordinates.")
        x, y = int(match.group(1)), int(match.group(2))
        approved, _ = self._maybe_approve("gui_click", "Gemini GUI click request", f"{x},{y}", False, {"x": x, "y": y})
        if not approved:
            return GatewayResult(False, "[BLOCKED] User denied GUI click.")
        try:
            import pyautogui
        except Exception as exc:
            return GatewayResult(False, f"[ERROR] pyautogui is not installed: {exc}")
        pyautogui.click(x, y)
        return GatewayResult(True, f"[OK] Clicked at {x},{y}")

    def _tool_gui_type(self, raw: str) -> GatewayResult:
        text = str(raw or "")
        approved, edited = self._maybe_approve("gui_type", "Gemini GUI type request", text, True, {"characters": len(text)})
        if not approved:
            return GatewayResult(False, "[BLOCKED] User denied GUI typing.")
        try:
            import pyautogui
        except Exception as exc:
            return GatewayResult(False, f"[ERROR] pyautogui is not installed: {exc}")
        pyautogui.write(edited, interval=0.01)
        return GatewayResult(True, f"[OK] Typed {len(edited)} characters.")

    def _tool_gui_hotkey(self, raw: str) -> GatewayResult:
        keys = [part.strip() for part in re.split(r"[,+\s]+", str(raw or "")) if part.strip()]
        if not keys:
            return GatewayResult(False, "[ERROR] gui_hotkey requires keys such as ctrl,shift,s.")
        approved, _ = self._maybe_approve("gui_hotkey", "Gemini GUI hotkey request", ",".join(keys), False, {"keys": ",".join(keys)})
        if not approved:
            return GatewayResult(False, "[BLOCKED] User denied GUI hotkey.")
        try:
            import pyautogui
        except Exception as exc:
            return GatewayResult(False, f"[ERROR] pyautogui is not installed: {exc}")
        pyautogui.hotkey(*keys)
        return GatewayResult(True, f"[OK] Sent hotkey: {','.join(keys)}")

    def _tool_blender_open(self, raw: str) -> GatewayResult:
        target = str(raw or "").strip()
        command = [self._blender_command()]
        metadata = {}
        if target:
            path = self._resolve_allowed_path(target, must_exist=True, access_reason="open Blender file")
            command.append(str(path))
            metadata["file"] = str(path)
        approved, _ = self._maybe_approve("blender_open", "Gemini Blender open request", " ".join(command), False, metadata)
        if not approved:
            return GatewayResult(False, "[BLOCKED] User denied Blender open.")
        try:
            subprocess.Popen(command, cwd=self._terminal_cwd())
        except FileNotFoundError:
            return GatewayResult(False, "[ERROR] Blender executable was not found. Set mcp_servers.blender.command in config.json.")
        return GatewayResult(True, f"[OK] Blender launch requested: {' '.join(command)}")

    def _tool_blender_python(self, raw: str) -> GatewayResult:
        script = str(raw or "").strip()
        if not script:
            return GatewayResult(False, "[ERROR] blender_python requires a Python script.")
        approved, edited = self._maybe_approve("blender_python", "Gemini Blender Python request", script, True, {"mode": "background"})
        if not approved:
            return GatewayResult(False, "[BLOCKED] User denied Blender Python.")
        with tempfile.NamedTemporaryFile("w", suffix=".py", delete=False, encoding="utf-8") as handle:
            handle.write(edited)
            script_path = handle.name
        try:
            return self._run_process([self._blender_command(), "--background", "--python", script_path], cwd=self._terminal_cwd(), timeout=120, label="BLENDER")
        except FileNotFoundError:
            return GatewayResult(False, "[ERROR] Blender executable was not found. Set mcp_servers.blender.command in config.json.")
        finally:
            try:
                os.unlink(script_path)
            except OSError:
                pass

    def _blender_command(self) -> str:
        return str((self.config.get("mcp_servers", {}).get("blender") or {}).get("command") or "blender")

    def _tool_excel_info(self, raw: str) -> GatewayResult:
        workbook, path = self._load_workbook_for_read(raw)
        try:
            payload = {
                "path": str(path),
                "sheets": [
                    {
                        "name": ws.title,
                        "max_row": ws.max_row,
                        "max_column": ws.max_column,
                        "dimensions": ws.calculate_dimension(),
                    }
                    for ws in workbook.worksheets
                ],
            }
            return GatewayResult(True, json.dumps(payload, indent=2, ensure_ascii=False))
        finally:
            workbook.close()

    def _tool_excel_read(self, raw: str) -> GatewayResult:
        data = parse_key_value_block(raw)
        workbook, path = self._load_workbook_for_read(data.get("path") or raw)
        sheet_name = data.get("sheet") or workbook.sheetnames[0]
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            return GatewayResult(False, f"[ERROR] Sheet not found: {sheet_name}")
        try:
            ws = workbook[sheet_name]
            cell_range = data.get("range") or ws.calculate_dimension()
            rows = []
            for row in ws[cell_range]:
                rows.append("\t".join("" if cell.value is None else str(cell.value) for cell in row))
            return GatewayResult(True, f"[EXCEL {path} | {sheet_name}!{cell_range}]\n" + "\n".join(rows))
        finally:
            workbook.close()

    def _tool_excel_write(self, raw: str) -> GatewayResult:
        data = parse_key_value_block(raw)
        path = self._resolve_allowed_path(data.get("path", ""), must_exist=False, access_reason="write Excel workbook")
        sheet_name = data.get("sheet") or "Sheet1"
        cell = data.get("cell") or "A1"
        value = data.get("value", "")
        approved, edited = self._maybe_approve("excel_write", "Gemini Excel write request", value, True, {"path": str(path), "sheet": sheet_name, "cell": cell})
        if not approved:
            return GatewayResult(False, "[BLOCKED] User denied Excel write.")
        workbook = self._open_or_create_workbook(path)
        ws = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)
        ws[cell] = edited
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)
        workbook.close()
        return GatewayResult(True, f"[OK] Wrote Excel cell {sheet_name}!{cell} in {path}")

    def _tool_excel_append_row(self, raw: str) -> GatewayResult:
        data = parse_key_value_block(raw)
        path = self._resolve_allowed_path(data.get("path", ""), must_exist=False, access_reason="append Excel row")
        sheet_name = data.get("sheet") or "Sheet1"
        values_text = data.get("values", "")
        try:
            values = json.loads(values_text) if values_text.strip().startswith("[") else [part.strip() for part in values_text.split("|")]
        except Exception:
            values = [part.strip() for part in values_text.split("|")]
        approved, edited = self._maybe_approve("excel_append_row", "Gemini Excel append row request", json.dumps(values, ensure_ascii=False), True, {"path": str(path), "sheet": sheet_name})
        if not approved:
            return GatewayResult(False, "[BLOCKED] User denied Excel append row.")
        try:
            values = json.loads(edited)
        except Exception:
            values = [part.strip() for part in edited.split("|")]
        workbook = self._open_or_create_workbook(path)
        ws = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)
        ws.append(values)
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)
        workbook.close()
        return GatewayResult(True, f"[OK] Appended Excel row to {sheet_name} in {path}")

    def _load_workbook_for_read(self, raw_path: str):
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise RuntimeError(f"openpyxl is not installed: {exc}") from exc
        path = self._resolve_allowed_path(raw_path, must_exist=True, access_reason="read Excel workbook")
        return load_workbook(path, read_only=True, data_only=True), path

    def _open_or_create_workbook(self, path: Path):
        try:
            from openpyxl import Workbook, load_workbook
        except Exception as exc:
            raise RuntimeError(f"openpyxl is not installed: {exc}") from exc
        if path.exists():
            return load_workbook(path)
        workbook = Workbook()
        return workbook

    def _tool_think(self, raw: str) -> GatewayResult:
        note_path = self.data_dir / "memory" / "thinking_log.jsonl"
        note_path.parent.mkdir(parents=True, exist_ok=True)
        with note_path.open("a", encoding="utf-8") as handle:
            handle.write(json.dumps({"note": str(raw or "").strip()}, ensure_ascii=False) + "\n")
        return GatewayResult(True, "[OK] Thinking note recorded locally.")

    def _load_memory(self) -> dict:
        if not self.memory_path.exists():
            return {}
        try:
            mtime = self.memory_path.stat().st_mtime
            if self._memory_cache is not None and self._memory_cache_mtime >= mtime:
                return self._memory_cache
            data = json.loads(self.memory_path.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                self._memory_cache = data
                self._memory_cache_mtime = mtime
                return data
            return {}
        except Exception:
            return {}

    @staticmethod
    def _strip_html(text: str) -> str:
        value = re.sub(r"(?is)<script.*?</script>", " ", str(text or ""))
        value = re.sub(r"(?is)<style.*?</style>", " ", value)
        value = re.sub(r"(?i)<br\s*/?>", "\n", value)
        value = re.sub(r"(?i)</(p|div|section|article|li|ul|ol|h[1-6]|tr)>", "\n", value)
        value = re.sub(r"(?s)<[^>]+>", " ", value)
        value = html.unescape(value)
        value = re.sub(r"[ \t]+\n", "\n", value)
        value = re.sub(r"\n{3,}", "\n\n", value)
        return value.strip()
